from __future__ import annotations

import io
import struct
import zlib
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook


def _make_png_bytes(width: int, height: int) -> bytes:
    def _chunk(chunk_type: bytes, data: bytes) -> bytes:
        c = chunk_type + data
        return struct.pack(">I", len(data)) + c + struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)

    sig = b"\x89PNG\r\n\x1a\n"
    ihdr = _chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0))
    raw = bytearray()
    for _ in range(height):
        raw.append(0)  # filter byte
        raw.extend(b"\xcc\xcc\xcc" * width)
    idat = _chunk(b"IDAT", zlib.compress(bytes(raw)))
    iend = _chunk(b"IEND", b"")
    return sig + ihdr + idat + iend


def create_sample_pdf(path: Path) -> None:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    pdf.cell(text="Quantitative Trading: A Practical Guide")
    pdf.ln(10)
    pdf.multi_cell(w=0, text="This document describes quantitative trading strategies used by hedge funds and institutional investors. Key concepts include statistical arbitrage, mean reversion, and momentum trading.")
    pdf.ln(5)
    pdf.multi_cell(w=0, text="Chapter 1: Introduction to Algorithmic Trading. Algorithmic trading uses computer programs to execute trading strategies at speeds and frequencies that are impossible for human traders. The practice has grown significantly since the 2000s, accounting for over 60% of equity trading volume in major markets.")
    pdf.ln(5)
    pdf.multi_cell(w=0, text="Chapter 2: Statistical Arbitrage. Statistical arbitrage involves identifying and exploiting pricing inefficiencies between related financial instruments. Pairs trading is a common strategy where two historically correlated assets are monitored for price divergence.")
    pdf.ln(5)
    pdf.multi_cell(w=0, text="Chapter 3: Risk Management. Proper risk management is essential for long-term trading success. Position sizing, stop-loss orders, and portfolio diversification help mitigate potential losses during adverse market conditions.")
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    pdf.cell(text="Appendix: Data Sources")
    pdf.ln(10)
    pdf.multi_cell(w=0, text="Common data sources for quantitative research include: market data feeds from exchanges, alternative data providers, SEC filings, and economic indicators. Historical data quality and survivorship bias are important considerations.")
    pdf.output(str(path))


def create_sample_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(["Date", "Ticker", "Side", "Quantity", "Price", "P&L"])
    ws.append(["2024-01-05", "AAPL", "BUY", 100, 185.50, 0])
    ws.append(["2024-01-08", "AAPL", "SELL", 100, 192.30, 680.00])
    ws.append(["2024-01-10", "MSFT", "BUY", 50, 415.20, 0])
    ws.append(["2024-01-15", "MSFT", "SELL", 50, 428.10, 645.00])
    ws.append(["2024-01-20", "GOOGL", "BUY", 30, 142.80, 0])
    ws.append(["2024-01-25", "GOOGL", "SELL", 30, 148.50, 171.00])
    ws.append(["2024-02-01", "TSLA", "BUY", 25, 198.60, 0])
    ws.append(["2024-02-10", "TSLA", "SELL", 25, 210.40, 295.00])
    ws.append(["2024-02-15", "AMZN", "BUY", 40, 178.90, 0])
    ws.append(["2024-02-28", "AMZN", "SELL", 40, 185.30, 256.00])
    wb.save(str(path))


def create_image_pdf(path: Path) -> None:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    png_bytes = _make_png_bytes(100, 100)
    pdf.image(io.BytesIO(png_bytes), x=50, y=50, w=100, h=100)
    pdf.output(str(path))


if __name__ == "__main__":
    out = Path(__file__).parent
    create_sample_pdf(out / "sample_text.pdf")
    create_image_pdf(out / "sample_image.pdf")
    create_sample_xlsx(out / "sample.xlsx")
    print(f"Fixtures created in {out}")
